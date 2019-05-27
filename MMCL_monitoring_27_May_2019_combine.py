from os import getcwd
from openpyxl import *
from threading import Thread
import os
from tkinter import *
from datetime import datetime
from tkinter import Tk,filedialog,Frame,StringVar,Button,Label
from time import time
from os import getcwd
from calendar import month_name
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
from tkinter.colorchooser import *
import win32com.client as win32
from win32com.client import Dispatch
from win32com.client import *
from functools import partial

header_color = "ff0000"
side_color = "00ff00"

class idd_class:

    def __init__(self):
        self.filename = "Open csv file "

    def upload(self):
        #self.filename = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        self.filename = filedialog.askopenfilename(title = "Select IDD raw file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.filename !="":
            var_1.set(ins_text_1.filename)

    def set_border(self, ws, cell_range,bd_sty="thin",font_weight=False,bg_color = "ffffff"):  # Function to set cell properties
        #bd_sty = "thin"
        bd_color = "000000"
        border = Border(
            left=Side(border_style=bd_sty, color=bd_color),
            right=Side(border_style=bd_sty, color=bd_color),
            top=Side(border_style=bd_sty, color=bd_color),
            bottom=Side(border_style=bd_sty, color=bd_color)
        )  # this indentation helps to comment unnecessary command line.
        font = Font(name="Times New Roman",size=11,bold=font_weight,color="000000")
        # set font bold,italic are boolen, font color etc
        alignment = Alignment(horizontal="center",vertical="center",wrapText=True)
        patternfill = PatternFill(start_color=bg_color,end_color=bg_color, fill_type="solid")  # set cell background color
        print("Cell range from cell property method ",cell_range)
        rows = ws.iter_rows(min_row=cell_range[0],min_col=cell_range[1],max_row=cell_range[2],max_col=cell_range[3])  # iter_rows function make a tuple of tuples of cell objects.
        for row in rows:
            for cell in row:
                cell.border = border        # set Border properties
                cell.font = font            # set font styles
                cell.alignment = alignment  # set alignment
                cell.fill = patternfill     # set background color.
        return ws

    def csv_file(self):
        # print(self.filename)
        #try:
        fd1 = open(self.filename, "r+");
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_idd = Workbook();
        sh_list = wb_idd.sheetnames;
        sh_list[0] ="idd raw";
        sh_list = wb_idd.sheetnames;
        ws_idd = wb_idd[sh_list[0]];
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A]*6 +[B]*3+[C]+[D,D,E,I]+ [P]*3+[S]*7;  #to be delete column list
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):    # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:         # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            row_list.insert(0," ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_connect_number = row_list.index("Connect Number")
                index_attempt_number = row_list.index("Attempt Number")
                index_Answer_Number  = row_list.index("Answer Number")
                index_Answer_Time    = row_list.index("Answer Time")
                row_list.append("ASR")
                row_list.append("ACD")
                row_list.append("CCR")
            else:
                max_ro = itm;
                row_list.append(f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60")
                row_list.append(f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")

            ws_idd.append(row_list);
        fd1.close();
        min_row = ws_idd.min_row
        min_col = ws_idd.min_column
        max_row = ws_idd.max_row
        max_col = ws_idd.max_column
        #col_letter = get_column_letter(max_col)

        ''''
        Setting style to the border and fonts, 
        '''
        cell_range = [min_row,min_col+1,max_row,max_col]  #set style all sheet.
        self.set_border(ws_idd,cell_range)

        top_cell_range = [1,2,1,ws_idd.max_column]  # set style top row/ Header row
        self.set_border(ws_idd,top_cell_range,bd_sty="medium",font_weight=True,bg_color=header_color)  # Border width = medium valid !!

        top_cell_range = [2,ws_idd.max_column-2,ws_idd.max_row,ws_idd.max_column]  # set style ASR,ACD,CCR
        self.set_border(ws_idd,top_cell_range,bd_sty="thin",font_weight=True,bg_color=side_color)

        rd = ws_idd.row_dimensions[1]  # get dimension for row 3
        rd.height = 48  # value in points, there is no "auto"
        #sd1 = [16,14,9,12,15,15,12,10,10,11,11,10M,12,11,11,13Q,11,11,11]
        sd1 = [4,16,14,9,12,15,15,12,10,10,11,11,10,12,11,11,11, 13,11,11,9,9,9]
        print(len(sd1))
        for sd in range(1,ws_idd.max_column+1):
            cd = ws_idd.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd-1]
        ws_idd.freeze_panes = "A2"  # make freeze before B2
        # ws_idd.insert_cols(1)  # this will insert correctly but little problem with my fix excel formulas

        #  floating number will show up to two decimal point
        num_f = ws_idd.iter_rows(min_row=2,min_col=ws_idd.max_column-2,max_row=ws_idd.max_row,max_col=ws_idd.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        # calculate upto time
        for itm in range(2,30):
            a = int(ws_idd.cell(itm,4).value)    # 'D' column contain Hours informations
            b = int(ws_idd.cell(itm+1,4).value)     # 'D' column contain Hours informations
            if a>b:
                c=a
                break
        time_upto = datetime.strptime(f"{c}", "%H").strftime("%I %p")
        print("last hours in 24H format ",c,time_upto)

        # save file
        date_file = (ws_idd["C5"].value).split("-")     # 'C' column contain date informations
        print(date_file)
        month_file = month_name[int(date_file[1])]
        name = f"IDD Report {date_file[2]} {month_file} {date_file[0]} (Every 2 Hours).xlsx"
        wb_idd.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        xl = Dispatch("Excel.Application")
        wb2 = xl.Workbooks.Open(Filename=path2)
        xl.Visible = False  # speed up process also
        wb2.Save()  # Save and over lap the original file
        wb2.Close(True)
        xl.Quit()

        print("IDD Report Done.\n","*"*20,"\n")
        # set GUI bottom Label textvariable . var_3 is a tkinter.StringVar() object
        # Print completion message
        var_3.set("IDD Report Done.\n" + "*"*20)
        # except Exception: print("File not found \n Or Wrong file selected \nOut file has opened")

        # mailing purpose only..
        idd_class.filename_mail_2h = name
        idd_class.filename_mail_2h_t = [date_file[0],month_file, date_file[2],f'{time_upto}']   # [ year, month, day , upto ]
        if auto_mail_enable.get() == 1:
            mail_instance.idd_every_two_hours()

        return None


class ccr_class:

    def __init__(self):
        self.filename = "Open csv file "

    def upload(self):
        self.filename = filedialog.askopenfilename(title = "Select file for CCR Check",filetypes = (("csv files","*.csv"),("all files","*.*")))
        #self.filename = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.filename !="":
            var1.set(ins_text.filename)

    def csv_file(self):
        instance_of_idd = idd_class();  # i will use it to set border property
        try:
            print(self.filename)
            fd1 = open(self.filename, "r+");   fd1.seek(0, 0);
            length_fd1 = len(fd1.readlines());
            fd1.seek(0, 0);
            wb_ccr = Workbook();
            # Dhaka = wb_in_kpi.create_sheet("Dhaka", 0)  # insert at first position
            sh_list = wb_ccr.sheetnames;
            ccr = wb_ccr[sh_list[0]];
            ccr.title = "CCR Check"
            fd1.seek(0, 0);
            countries = ["India", "Malaysia", "Saudi Arabia", "Singapore", "UAE"]
            A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
            del_col = [0]*6 +[10]*9 + [18]*13;
            print("CCR column Delete list : ",del_col);
            fd1.seek(0, 0); #this command take file pointer at 0,0 position
            for itm in range(length_fd1): #this loop work at the end of the line  of csv file
                row_list = fd1.readline().split('","');
                for item in del_col:  # delete unnecessary column from .csv file
                    del row_list[item];
                for item1 in range(len(row_list)):  #convert str to int
                    if row_list[item1].isdecimal():
                        row_list[item1] = int(row_list[item1])
                if itm == 0:
                    row_list[len(row_list) - 1] = "CCR";  # for the first row put CCR at the top
                    index_connect_number = row_list.index("Connect Number")  # find index of connect number to calculate ccr
                    index_attempt_number = row_list.index("Attempt Number")
                    index_CURRENT_TIME = row_list.index("CURRENT TIME")  # to filter time, this is index of a list not xlsx
                    # print("connect and attempt numbers = ",index_connect_number,index_attempt_number)
                else:
                    try:
                        # row_list[len(row_list)-1] = int(row_list[index_connect_number])/int(row_list[index_attempt_number])*100;
                        row_list[len(row_list)-1] = int(row_list[index_connect_number])/int(row_list[index_attempt_number])*100;
                    except ZeroDivisionError:pass
                        #print("zero division Error ")
                    except Exception: print("Unknown Error ")
                if row_list[0] in countries:  # skip unnecessary countries
                    max_ro = ccr.max_row
                    row_list.append(f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                    ccr.append(row_list)
                elif row_list[0] == "Object Description":
                    row_list.append("CCR")
                    ccr.append(row_list);
            fd1.close()
            ccr.auto_filter.ref = f"A1:Z{ccr.max_column}"
            ccr.auto_filter.add_filter_column(0, ["India", "Malaysia", "Saudi Arabia","Singapore","UAE"])
            time_slot = [];
            time_slot_2 = [];
            time_slot_range = []
            for fil_1 in range(2,30):
                var23 = ccr.cell(fil_1, index_CURRENT_TIME+1).value # since list index start with 0,
                # and xlsx index start with 1 that's why here add 1 (one).
                time_slot_2.append(var23)
            print("time slot_2 = ",time_slot_2)
            for fil_3 in range(27):
                time_slot.append(int(time_slot_2[fil_3][:2]))
            print("time slot = ", time_slot)
            for fil_2 in range(24):
                if time_slot[fil_2]>time_slot[fil_2+1]:
                    time_slot_range.append(time_slot_2[fil_2-2])
                    time_slot_range.append(time_slot_2[fil_2-1])
                    time_slot_range.append(time_slot_2[fil_2])
                    break;
            print("Last three hours: ", time_slot_range)
            ccr.auto_filter.add_filter_column(index_CURRENT_TIME, time_slot_range)

            # set border and cell property to Header cells
            instance_of_idd.set_border(ccr,[1,1,1,ccr.max_column],bd_sty="medium",font_weight=False,bg_color=header_color)
            # set border to data cells
            instance_of_idd.set_border(ccr,[2,1,ccr.max_row,ccr.max_column])
            # set font color Red and weight bold it CCR less than 80%
            low_ccr_list = []
            for color1 in ccr.iter_rows(min_row=3, min_col=1):  # red mark ccr less than 80%
                if int(color1[ccr.max_column - 2].value) < 80:
                    for color2 in color1:
                        color2.font = Font(bold=False, color="ff0000")
                        low_ccr_list.append(color2.value)
            print(f"Low CCR count today = {len(low_ccr_list)/ccr.max_column}")

            sd1 = [16, 14, 25, 16, 12, 14, 11, 11, 10, 12, 13, 12, 12, 11, 13, 11, 13, 11, 11, 9, 9, 9]
            print(len(sd1))
            for sd in range(1, ccr.max_column + 1):
                cd = ccr.column_dimensions[f"{get_column_letter(sd)}"]
                cd.width = sd1[sd - 1]
            rd = ccr.row_dimensions[1]  # get dimension for row 3
            rd.height = 45  # value in points, there is no "auto"
            ccr.freeze_panes = "E2"  # make freeze before E2

            #  floating number will show up to two decimal point
            num_f = ccr.iter_rows(min_row=2,min_col=ccr.max_column-1,max_row=ccr.max_row,max_col=ccr.max_column)
            for num_1 in num_f:
                for num_2 in num_1:
                    num_2.number_format = '#,##0.00'

            ccr_out_file = f"CCR_ISD_Check up to {time_slot_range[2]}.xlsx"
            ccr_out_file = ccr_out_file.replace(":","_")  # file name do not contain : sign
            print("Out file name ccr : ",ccr_out_file)
            wb_ccr.save(ccr_out_file);

            # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
            path1 = os.getcwd() + os.sep
            path2 = os.path.join(path1, ccr_out_file)
            xl = Dispatch("Excel.Application")
            wb2 = xl.Workbooks.Open(Filename=path2)
            xl.Visible = False  # speed up process also
            wb2.Save()  # Save and over lap the original file
            wb2.Close(True)
            xl.Quit()

            # set GUI bottom Label textvariable . var_3 is a tkinter.StringVar() object
            # Print completion message
            var_3.set("CCR Check file has Done.\n"+"*"*20)
            print("CCR Check file has Done.\n","*"*20,"\n")
        except Exception:print("file not found or Wrong file selected or Out file has Opened")


class upload:

    # make class variable
    wb_kpi_raw_in = Workbook()  # not used yet
    wb_kpi_raw_out = Workbook()
    filename = ' Open File .. .. .. .'
    all_city_trunk = [];
    ver_sum = 0;
    def __init__(self):
        self.wb_raw_data = Workbook()
        self.ws_raw_data = self.wb_raw_data[self.wb_raw_data.sheetnames[0]]
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem)
        upload.index_of_answer_time = 0 # this is a class variable. same memory point for all functions
        self.total_time = 0
        #upload.wb_kpi_in = Workbook()  # this is a class variable. same memory point for all functions

    # unused function
    def fileupload(self):
        self.filename = filedialog.askopenfilename();
        if self.filename[-4:] == ".xls":
            self.open_xls_as_xlsx();
        elif self.filename[-4:] == ".csv":
            self.csv_to_xlsx_in_kpi();
        if self.filename == '':
            self.filename=" Open file .. .. ."
            print(type(self.filename))
        raw_in_kpi_file_path.set(raw_kpi_in_instance.filename)
        raw_out_kpi_file_path.set(raw_kpi_out_instance.filename)
        raw_idd_file_path.set(raw_idd.filename)
        # sample_in_kpi_file_path.set(sample_kpi_report.filename)
        # sample_IOS_report_path.set(sample_isd_report.filename)
        # sample_idd_file_path.set(sample_idd_report.filename)
        return None

    def upload_kpi(self):
        self.filename = filedialog.askopenfilename(title = "Select incoming KPI raw CSV file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.filename != '':
            raw_in_kpi_file_path.set(raw_kpi_in_instance.filename)
            self.uploaded_cvs_to_xlsx_kpi()
            #raw_kpi_in_instance.uploaded_cvs_to_xlsx_kpi()

    def upload_kpi_out(self):
        self.filename = filedialog.askopenfilename(title = "Select outgoing KPI Raw CSV File",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.filename != '':
            raw_out_kpi_file_path.set(raw_kpi_out_instance.filename)
            raw_kpi_out_instance.uploaded_cvs_to_xlsx_kpi()
        return self.filename

    def upload_idd(self):
        self.filename = filedialog.askopenfilename(title = "Select idd Raw CSV File",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.filename != '':
            raw_idd_file_path.set(self.filename)
        return self.filename

    def uploaded_cvs_to_xlsx_kpi(self):
        fd1 = open(self.filename, "r+");
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_kpi = Workbook();
        sh_list = wb_kpi.sheetnames;
        ws_kpi = wb_kpi[sh_list[0]]
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A]*6 +[B]*3+[C,C]+[F,G,K]+ [R]*3+[U,U]+[V]*4;  #to be delete column list
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):    # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:         # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            # row_list.insert(0," ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_Answer_Time    = row_list.index("Answer Time")
                upload.index_of_answer_time = index_Answer_Time   # this is a class variable. same memory point for all functions

            ws_kpi.append(row_list);
        fd1.close()
        self.wb_raw_data = wb_kpi;
        self.ws_raw_data = wb_kpi[wb_kpi.sheetnames[0]]
        # wb_kpi.save("_in_kpi_cut file.xlsx")  # save data only for testing purpose
        # os.chdir(os.getcwd()+ "\ new")
        return ws_kpi

    def trunk_finder(self,raw_kpi_ws):
        list1 = []  # Local variable
        list2 = []
        for itm in range(1, raw_kpi_ws.max_row):
            if raw_kpi_ws.cell(itm, 1).value not in list1:
                list1.append(raw_kpi_ws.cell(itm, 1).value)
                list2.append(itm)
        return list1, dict(zip(list1, list2))

    # useless function
    def object_description_finder(self,raw_kpi_ws):
        list_object_description = []  # Local variable
        for itm in range(1, raw_kpi_ws.max_column):
            if raw_kpi_ws.cell(1, itm).value not in list_object_description:
                list_object_description.append(raw_kpi_ws.cell(itm, 1).value)
        return list_object_description

    def city_trunk_finder(self,list_kpi_trunk, key):
        list_kpi_trunk_city = []
        for itm in list_kpi_trunk:
            #print("dictionary elements ",itm)
            if not key != itm[-2:]:  # for Dhaka key = DH; khulna key=KH;
                sum = self.answer_time_sum(itm)
                if sum>0:
                    list_kpi_trunk_city.append(itm)
        # print("list_kpi_trunk_city ",list_kpi_trunk_city)
        return list_kpi_trunk_city

    def city_report(self,city,list_kpi_trunk_city,ws_raw_data,dic_trunk_col_raw,time_range):
        # Creation of city Worksheet instance
        if f"{city}" in self.wb_report.sheetnames:
            temp_ws = self.wb_report[f"{city}"]
            self.wb_report.remove(temp_ws)  # remove sheet if it already exist
            self.wb_report.create_sheet(f"{city}")  # a new worksheet will creates with city name title
        else:
            self.wb_report.create_sheet(f"{city}")  # a new worksheet will creates with city name title
        City = self.wb_report[f"{city}"]  # a new worksheet will be in City variable.
        # City.title(f"{city}") # value less code. title already given when created.!!

        # First row of the city report: Header of city report.
        header_row = [ws_raw_data.cell(1,hr).value for hr in range(1,self.ws_raw_data.max_column+1)] # list com
        header_row.insert(0," ")
        header_row.append("ASR")
        header_row.append("ACD")
        header_row.append("CCR")
        header_row[3] = "Start Time"    # "Start Time" is more meaningful than "Last Time"
        header_row[4] = "End Time"      # "End Time" is more meaningful than "End Time"
        for itm,itm1 in zip(header_row,range(1,1+len(header_row))): City.cell(1,itm1).value = itm
        #City.append(header_row)
        cell_range_h = [1,2,1,len(header_row)]
        self.set_border(City, cell_range=cell_range_h,bd_sty="medium",font_weight=True,bg_color=header_color)

        index_connect_number = header_row.index("Connect Number")
        index_attempt_number = header_row.index("Attempt Number")
        index_Answer_Number = header_row.index("Answer Number")
        index_Answer_Time = header_row.index("Answer Time")

        # self.wb_report.save("test_report.xlsx")

        #print(dic_trunk_col_raw)
        length = len(list_kpi_trunk_city)
        # print("length of city list ******** ", length,list(range(length)))
        # print("city list ",list_kpi_trunk_city)

        list1 = {}

        # for itm, item in list_kpi_trunk_city,range(length):  # nor working
        for itm, item in zip(list_kpi_trunk_city,range(length)):  # for all trunk in a city.
            # range used only for numaric variable to use. useless !!
            start_row = dic_trunk_col_raw.get(itm)  # this will give the starting point of the data
            for itm1 in range(1,time_range+1):
                for itm2 in range(1,ws_raw_data.max_column+1):
                    row = start_row + (itm1-1)
                    col = itm2
                    temp1 = ws_raw_data.cell(row, col).value;  # collect value form raw file.
                    row2 = time_range*item + (itm1-1)+2*(item+1)  #
                    col2 = itm2 +1 # plus 1 for report start with second column
                    City.cell(row2, col2).value = temp1;  # put value in
                max_ro = itm1;
                City.cell(row2, col2 + 1).value = f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100"
                City.cell(row2, col2 + 2).value = f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60"
                City.cell(row2, col2 + 3).value = f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100"
            list1[f"{itm}"] = f"L{row2-time_range+1}:L{row2}"  # plus +1 means -> from 10 to 20 = 11
            City.merge_cells(start_row=row2-time_range+1, start_column=2, end_row=row2,end_column=2)
            City.merge_cells(start_row=row2-time_range+1, start_column=3, end_row=row2,end_column=3)

            # apply style to the sheets
            cell_range = [row2-time_range+1,2,row2,col2]   # apply to all values
            self.set_border(City,cell_range=cell_range,wraptext=False)  # only this function need 2.5 seconds to set border !!

            # merged cell must be bold. . .
            cell_range = [row2 - time_range + 1, 2, row2, 3]
            self.set_border(City, cell_range=cell_range, font_weight=True)

            # set side/ACD,CCR cell styles
            cell_range_cal = [row2-time_range+1,col2 + 1,row2,col2 + 3]    # apply to ASR, ASD, CCR columns
            self.set_border(City,cell_range=cell_range_cal,bd_sty="medium",bg_color=side_color,wraptext=False)

        # set Column width
        sd1 = [1, 29, 12, 8, 8, 9, 11, 11, 12, 12, 9, 11, 9, 9, 9, 9, 9, 9, 10, 12, 10, 10,11,8,8,8,10]
        # print(len(sd1))
        for sd in range(1, City.max_column + 1):
            cd = City.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        rd = City.row_dimensions[1]  # get dimension for row 1
        rd.height = 45  # value in points, there is no "auto"
        City.freeze_panes = "A2"  # make freeze before A2, means up to A1,B1,C1 . . .. .

        #  floating number will show up to two decimal point
        num_f = City.iter_rows(min_row=2, min_col=City.max_column - 2, max_row=City.max_row, max_col=City.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        return list1

    def make_kpi_report(self):
        # renew self.wb_report at first, if accidently click double on make report
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem) # remove the atumatic created first Sheet

        self.total_time = 0 # Total domestic Answer time set 0
        time1=time.time()
        kpi_path = raw_kpi_in_instance.filename;
        # print(kpi_path)
        sheet_names = self.wb_raw_data.sheetnames;

        raw_kpi_ws = self.wb_raw_data[sheet_names[0]]
        # print(raw_kpi_ws.max_row)
        list_kpi_trunk, self.dic_trunk_col_raw = self.trunk_finder(raw_kpi_ws)  # trunk and row in raw file.
        # print(self.dic_trunk_col_raw)
        # exit()
        time_range = self.time_range_cal(raw_kpi_in_instance.ws_raw_data)  # time range means up to XX:XX PM

        # list_object_description = self.object_description_finder(raw_kpi_ws)
        list_city_kpi_trunk=[]      # only trunks which have Answer Time value of voice call.
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "DH"))
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "CH"))
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "KH"))
        self.all_city_trunk = [list_city_kpi_trunk[0],list_city_kpi_trunk[1],list_city_kpi_trunk[2]]
        # insert Value for Dhaka
        # self.kpi_report = Workbook();

        dic_dh = self.city_report(city="Dhaka",list_kpi_trunk_city=list_city_kpi_trunk[0],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)
        dic_ch = self.city_report(city="CTG",list_kpi_trunk_city=list_city_kpi_trunk[1],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)
        dic_kh = self.city_report(city="Khulna",list_kpi_trunk_city=list_city_kpi_trunk[2],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)

        all_dic = [dic_dh,dic_ch,dic_kh]
        self.summary(all_dic)
        # print(all_dic)
        # print("dic_for summary second",all_dic)

        # save file with name
        x = [0,0]
        for itm in range(2,30):
            x[0] = int(self.ws_raw_data.cell(itm,3).value.split(":")[0])
            x[1] = int(self.ws_raw_data.cell(itm+1,3).value.split(":")[0])
            if x[0]>x[1]:
                break
        # second row and second column has date.
        date_file = self.ws_raw_data.cell(2, 2).value.split("-")
        month_file = month_name[int(date_file[1])]
        if x[0] < 23:
            time_upto = datetime.strptime(f"{x[0]+1}", "%H").strftime("%I %p")
            name = f"ICX_KPI Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"ICX_KPI Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''
        self.wb_report.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        xl = Dispatch("Excel.Application")
        wb2 = xl.Workbooks.Open(Filename=path2)
        xl.Visible = False  # speed up process also
        wb2.Save()  # Save and over lap the original file
        wb2.Close(True)
        xl.Quit()

        # Mailing purpose
        upload.nam_temp_kpi = name
        upload.nam_temp_kpi_t = [date_file[0], month_file, date_file[2], time_upto]
        if auto_mail_enable.get() == 1:
            mail_instance.kpi()

        print("Total domestic Answer time = ", self.total_time/60)
        print("file save successfully !! ")

        time2=time.time()
        print("KPI report done, time = ",time2-time1)
        kip_conform.set(f" KPI report done \nTime needed: {time2-time1}  Total Minutes: {self.total_time/60}")
        return None

    def make_IOS_ISD_report(self):
        # renew self.wb_report at first, if accidently click double on make report
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem)     # remove the atumatic created first Sheet

        self.total_time = 0  # Total domestic Answer time set 0
        time1 = time.time()
        sheet_names = self.wb_raw_data.sheetnames;

        raw_kpi_out_ws = self.wb_raw_data[sheet_names[0]]
        # print(raw_kpi_out_ws.max_row)
        list_kpi_trunk, self.dic_trunk_col_raw = self.trunk_finder(raw_kpi_out_ws)  # trunk and row in raw file.
        # print(self.dic_trunk_col_raw)

        timeR = self.time_range_cal(raw_kpi_out_ws)  # time range means up to XX:XX PM
        timeR1 = self.time_range_cal(raw_kpi_in_instance.ws_raw_data)  # time range means up to XX:XX PM

        # report work as IOS report.
        IOS_trunk_list = [["1001-Roots_IOS_IN"],["1005-NovoTel_IOS_IN"],["1009-Btrac_IOS_IN"],["1013-MirTelecom_IOS_IN"],["1017-Global_Voice_IN"],["1021-Unique_IOS"],["1025-Digicon_IOS"]]
        IOS_list = ["Roots","NovoTel","Btrac","MirTelecom","GlobalVoice","Unique","Digicon"]
        # IOS_trunk_list.reverse(); IOS_list.reverse()
        dic_ios = [] # list of trunk and value of excel ranges
        for itm in range(len(IOS_list)):
            dic_ios.append(self.city_report(city=f"{IOS_list[itm]}", list_kpi_trunk_city=IOS_trunk_list[itm],ws_raw_data=raw_kpi_in_instance.ws_raw_data,dic_trunk_col_raw=raw_kpi_in_instance.dic_trunk_col_raw,time_range = timeR1))
        # print("dic_iso = ",dic_ios)

        #   Summation of IOS ISD calls by function
        self.summary_ios(dic_ios)

        #   ISD report
        ans_isd_trunk = [["1050-TeleTalk_ISD"],["1070-GP1_ISD","1072-GP2_ISD"],["1080-Robi_ISD"],["1090-Banglalink_ISD"]]
        ans_isd_list = ["TaleTalk ISD","GP ISD","ROBI ISD","Banglalink ISD"]
        # ans_isd_list.reverse();  ans_isd_trunk.reverse();
        for itm in range(len(ans_isd_trunk)):
            self.city_report(city=f"{ans_isd_list[itm]}", list_kpi_trunk_city=ans_isd_trunk[itm],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range = timeR)

        # save file with name
        x = [0, 0]
        for itm in range(2, 30):
            x[0] = int(self.ws_raw_data.cell(itm, 3).value.split(":")[0])
            x[1] = int(self.ws_raw_data.cell(itm + 1, 3).value.split(":")[0])
            if x[0] > x[1]:
                break
        # second row and second column has date.
        date_file = self.ws_raw_data.cell(2, 2).value.split("-")
        month_file = month_name[int(date_file[1])]
        if x[0] < 23:
            time_upto = datetime.strptime(f"{x[0] + 1}", "%H").strftime("%I %p")
            name = f"IOS & ISD Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"IOS & ISD Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''
        self.wb_report.save(name)
        print("Total domestic Answer time = ", self.total_time / 60)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        xl = Dispatch("Excel.Application")
        wb2 = xl.Workbooks.Open(Filename=path2)
        xl.Visible = False  # speed up process also
        wb2.Save()  # Save and over lap the original file
        wb2.Close(True)
        xl.Quit()

        time2 = time.time()
        print("KPI report done, time = ", time2 - time1)

        # Calculate total Answer time for outgoing kpi file
        ans_isd_trunk1 = ["1050-TeleTalk_ISD","1070-GP1_ISD","1072-GP2_ISD","1080-Robi_ISD","1090-Banglalink_ISD"]
        for itm in ans_isd_trunk1: self.answer_time_sum(itm)
        kip_conform.set(f" IOS IDD report done \n Time needed: {time2-time1} Total Minutes : {self.total_time/60} (outgoing kpi)")

        # for Mailing purpose
        upload.nam_temp_ios = name
        upload.nam_temp_ios_t = [date_file[0], month_file, date_file[2],time_upto]
        if auto_mail_enable.get() == 1:
            mail_instance.ios_idd()

        return None

    # useless function
    def verification_sum1(self,dic_trunk_col_raw,raw_kpi_ws):
        self.ver_sum = 0;
        print("self.all_city_trunk = ", self.all_city_trunk,"\n length = ",len(self.all_city_trunk))
        for itm in self.all_city_trunk:
            #tmp = raw_kpi_ws.cell(itm, 2).value
            print(itm)
            tmp2 = dic_trunk_col_raw.get(itm)
            for item in range(24):
                self.ver_sum = self.ver_sum + int(raw_kpi_ws.cell(tmp2+item, 11).value)
        print("Total sum = ", self.ver_sum/60)

    def answer_time_sum(self,trunk):
        row = self.dic_trunk_col_raw.get(trunk)
        ws = self.wb_raw_data[self.wb_raw_data.sheetnames[0]]
        a = upload.index_of_answer_time+1
        b=self.time_range_cal(self.ws_raw_data)
        sum1 = 0
        for itm in range(b):
            sum1 += ws.cell(row+itm,a).value
        print("summission of a trunk: ",trunk,sum1)
        self.total_time +=sum1
        return sum1

    def time_range_cal(self,ws_raw_data):
        print("this is time range function")
        last_time = 3 # just random value
        for itm1 in range(1,10):
            if ws_raw_data.cell(1, itm1).value == "LAST TIME":
                last_time = itm1
                print("last time found : ",last_time)
                break
        list1 = []
        list2 = []
        for itm2 in range(2,40):
            list1.append(ws_raw_data.cell(itm2,last_time).value)
        for itm in list1:
            list2.append(int(itm[:2]))
        for itm3 in range(30):
            if list2[itm3]>list2[itm3+1]:
                time_range = list2[itm3]+1
                break
        self.time_range = time_range
        return time_range

    def summary(self, all_dic):
        # remove sheet if exist already
        if "Summary" not in self.wb_report.sheetnames:
            self.wb_report.create_sheet(title="Summary",index=3)
        else:
            temp_ws = self.wb_report["Summary"]
            self.wb_report.remove(temp_ws)
            self.wb_report.create_sheet(title="Summary",index=3)
        ws_summary = self.wb_report["Summary"]

        #Border Style set
        bd_sty = "medium"
        font_weight = True

        # all_dic = list of dictionary of city trunks with value "Answer Time" ranges it will help to make sum.
        list1 =["Zone","Trunk ID","Total Time (Sec)","Total Time (Min)"]
        list2 = [["Zone","Total Domestic Time (Min)"],["Dhaka",0],["CTG",0],["Khulna",0]]

        # self.ws_summary.cell(2,7).value = list1

        # variables that used
        col_ind = 9     # How many column will be empty before summary calculation
        len0 = 3        # position of Dhaka trunk summary
        len1 = len(self.all_city_trunk[0]) + len0 + 4   # CTG trunk summary starts
        len2 = len(self.all_city_trunk[1]) + len1 + 4
        len3 = len(self.all_city_trunk[2]) + len2 + 4
        length = [len0,len1,len2,len3]
        #make class variable
        upload.check_loop = 0

        # =SUM(Dhaka!L2: L25)
        for itm,itm2 in zip(list1,range(len(list1))):   # print header of each summary
            ws_summary.cell(len0, col_ind+itm2).value = itm  # print header of each summary
            ws_summary.cell(len1, col_ind+itm2).value = itm
            ws_summary.cell(len2, col_ind+itm2).value = itm
        # Dhaka  # Starting column -3 means first col of summary header
        cell_range_dhH = [len0,(col_ind+itm2)-3,len0,col_ind+itm2]
        # CTG
        cell_range_chH = [len1, (col_ind + itm2) - 3, len1, col_ind + itm2]
        # Khulna
        cell_range_khH = [len2, (col_ind + itm2) - 3, len2, col_ind + itm2]

        # ******************************************************************************************
        for var1, var2 in zip(self.all_city_trunk[0], range(len(self.all_city_trunk[0]))):
            upload.check_loop += 1
            list_dh = ["Dhaka", var1, f"=SUM(Dhaka!{all_dic[0].get(var1)})", f"=SUM(Dhaka!{all_dic[0].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len0+1) + var2, var4 + col_ind).value = var3
        # ws_summary.merge_cells("B1:B5") # this syntax also valid.
        ws_summary.merge_cells(start_row=(len0+1), start_column=col_ind, end_row=(len0+1) + var2, end_column=col_ind)
        sum_col = get_column_letter(var4 + col_ind)
        ws_summary.cell((len0+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len0+1)}:{sum_col}{(len0+1) + var2})"  # make summary of same city trunks in Minutes
        list2[1][1] = f"={sum_col}{(len0+1) + var2+1}"
        # cell ranges for simple style set
        cell_range_dh1 = [(len0+1), col_ind, (len0+1) + var2, var4 + col_ind]
        # so sad this function calls for only one cell !!
        cell_range_dh2 = [(len0 + 1)+var2+1, var4 + col_ind, (len0 + 1) + var2+1, var4 + col_ind]

        for var1, var2 in zip(self.all_city_trunk[1], range(len(self.all_city_trunk[1]))):
            list_dh = ["CTG", var1, f"=SUM(CTG!{all_dic[1].get(var1)})",
                       f"=SUM(CTG!{all_dic[1].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len1 + 1) + var2, var4 + col_ind).value = var3
        ws_summary.merge_cells(start_row=(len1 + 1), start_column=col_ind, end_row=(len1 + 1) + var2, end_column=col_ind)
        ws_summary.cell((len1+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len1+1)}:{sum_col}{(len1+1) + var2})"  # make summary of same city trunks in Minutes
        list2[2][1] = f"={sum_col}{(len1+1) + var2+1}"
        cell_range_ch1 = [(len1 + 1), col_ind, (len1 + 1) + var2, var4 + col_ind]
        cell_range_ch2 = [(len1 + 1) + var2+1,  var4 + col_ind, (len1 + 1) + var2+1 , var4 + col_ind]

        for var1, var2 in zip(self.all_city_trunk[2], range(len(self.all_city_trunk[2]))):
            list_dh = ["Khulna", var1, f"=SUM(Khulna!{all_dic[2].get(var1)})",
                       f"=SUM(Khulna!{all_dic[2].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len2 + 1) + var2, var4 + col_ind).value = var3
        ws_summary.merge_cells(start_row=(len2 + 1), start_column=col_ind, end_row=(len2 + 1) + var2, end_column=col_ind)
        ws_summary.cell((len2+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len2+1)}:{sum_col}{(len2+1) + var2})"  # make summary of same city trunks in Minutes
        list2[3][1] = f"={sum_col}{(len2+1) + var2+1}"
        # simple border style set ranges
        cell_range_kh1 = [(len2 + 1), col_ind, (len2 + 1) + var2, var4 + col_ind]
        cell_range_kh2 = [(len2 + 1)+var2+1, var4 + col_ind, (len2 + 1) + var2+1, var4 + col_ind]

        # Total Domestic Time (Min) summation
        for itm,va1 in zip(list2,range(4)):
            for itm1,va2 in zip(itm,range(2)):
                ws_summary.cell(len3+va1, (col_ind + 2)+va2).value = itm1
        col_let = get_column_letter((col_ind + 2)+va2)
        ws_summary.cell(len3 + va1 + 1, (col_ind + 2) + va2).value = f"=SUM({col_let}{len3+va1-2}:{col_let}{len3+va1})"
        # final summary cells border styles ranges
        cell_range_su1 = [len3+1, (col_ind+2), len3 + va1, va2 + (col_ind+2)]
        cell_range_su2 = [len3+va1+1, va2+(col_ind+2), len3+va1+1, va2 + (col_ind+2)]
        cell_range_suH = [len3, (col_ind+2), len3, col_ind + 3]  # Starting column -3 means first col of summary header

        # set all border style to all
        list_bd_sty_H = [cell_range_dhH,cell_range_chH,cell_range_khH,cell_range_suH]   # set Header style of summary
        list_bd_sty = [cell_range_dh1,cell_range_ch1,cell_range_kh1,cell_range_su1]     # range of cell of data
        list_bd_sty2 = [cell_range_dh2,cell_range_ch2,cell_range_kh2,cell_range_su2]

        for bd_sty_r in list_bd_sty: self.set_border(ws_summary,cell_range=bd_sty_r,wraptext=False,font_size=14)
        for bd_sty_r in list_bd_sty2: self.set_border(ws_summary,cell_range=bd_sty_r,wraptext=False,font_size=14,font_weight=True)
        for bd_sty_r in list_bd_sty_H: self.set_border(ws_summary,cell_range=bd_sty_r,bd_sty=bd_sty,bg_color=header_color,font_weight=font_weight,font_size=14)

        # set Column width
        sd1 = [9]*(col_ind-1)+[18, 34, 32, 33, 9,9,9,9,9,9]
        for sd in range(1,ws_summary.max_column + 1):
            cd = ws_summary.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        # set height of the headers
        for itm in length: ws_summary.row_dimensions[itm].height = float(23.75)
        #  floating number will show up to two decimal point
        num_f = ws_summary.iter_rows(min_row=2, min_col=col_ind+3, max_row=ws_summary.max_row, max_col=ws_summary.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'
        return None

    def summary_ios(self,all_dic):

        iso_summary = "IOS Summary"
        # remove sheet if exist already
        if iso_summary not in self.wb_report.sheetnames:
            self.wb_report.create_sheet(title=f"{iso_summary}",index=7)
        else:
            temp_ws = self.wb_report[iso_summary]
            self.wb_report.remove(temp_ws)
            self.wb_report.create_sheet(title=iso_summary,index=7)
        ws_summary = self.wb_report[iso_summary]

        # all_dic = list of dictionary of city trunks with value "Answer Time" ranges it will help to make sum.
        ios_trunk_list = ["1001-Roots_IOS_IN", "1005-NovoTel_IOS_IN", "1009-Btrac_IOS_IN", "1013-MirTelecom_IOS_IN", "1017-Global_Voice_IN", "1021-Unique_IOS","1025-Digicon_IOS"]
        ios_list = ["Roots", "NovoTel", "Btrac", "MirTelecom", "GlobalVoice", "Unique", "Digicon"]

        list1 =["IOS","Total Time (Sec)","Total Time (Min)"]
        list2 = ["IOS","Total International (Min)"]

        # variables that used
        col_ind = 9     # How many column will be empty before summary calculation
        len0 = 3        # position of Dhaka trunk summary
        len1 = len(ios_list) + len0 + 4   # CTG trunk summary starts
        length = [len0,len1]

        # Two hearer in excel
        for itm,itm2 in zip(list1,range(len(list1))):   # print header of each summary
            ws_summary.cell(len0, col_ind+itm2).value = itm  # print header of each summary
        for itm,itm2 in zip(list2,range(len(list2))):
            ws_summary.cell(len1, col_ind+itm2+1).value = itm

        #  set border style of header
        # Starting column -2 means first col of summary Header
        cell_range = []
        cell_range.append([len0,col_ind, len0,col_ind+2])
        # Minutes calculation
        cell_range.append([len1, (col_ind + 1), len1, col_ind + 2])
        for itm in range(2): self.set_border(ws_summary,cell_range[itm],bd_sty="medium",font_weight=True,bg_color=header_color,font_size=14)

        # ******************************************************************************************

        aa = all_dic[0].get(ios_trunk_list[0])
        print(aa)
        # exit()

        for var1, var2 in zip(ios_list, range(len(ios_list))):
            list_dh = [var1, f"=SUM({var1}!{all_dic[var2].get(ios_trunk_list[var2])})", f"=SUM({var1}!{all_dic[var2].get(ios_trunk_list[var2])})/60"]

            for var3, var4 in zip(list_dh, range(3)):
                ws_summary.cell((len0+1) + var2, var4 + col_ind).value = var3

            cell_range = [(len0+1) + var2, col_ind,(len0+1) + var2, var4 + col_ind]
            self.set_border(ws_summary,cell_range,font_size=14,wraptext=False)
        cell_range = [(len0 + 1), col_ind, (len0 + 1) + var2, col_ind]  # Cell range of IOS name to make font weight bold
        self.set_border(ws_summary,cell_range,wraptext=False,font_size=14,font_weight=True)

        # Total isd Time (Min) summation ************************
        for itm,va1 in zip(ios_list,range(len(ios_list))):
            sum_list = [f"{itm}",   f"=SUM({itm}!{all_dic[va1].get(ios_trunk_list[va1])})/60"]

            for itm1,va2 in zip(sum_list,range(2)):
                ws_summary.cell(len1+va1+1, (col_ind + 1)+va2).value = itm1

            # set all cell property/style
            cell_range = [len1+va1+1, (col_ind + 1), len1+va1+1, (col_ind + 1) + va2]
            self.set_border(ws_summary,cell_range,wraptext=False,font_size=14)

        col_let = get_column_letter((col_ind + 1)+va2)      # for sum calculation
        ws_summary.cell(len1 + va1 + 2, (col_ind + 1) + va2).value = f"=SUM({col_let}{len1+1}:{col_let}{len1+va1+1})"
        # print(f"=SUM({col_let}{len1+1}:{col_let}{len1+va1+1})")

        #   set final sum only one cell Style
        cell_range = [len1 + va1 + 2, (col_ind + 1) + va2, len1 + va1 + 2, (col_ind + 1) + va2]      # for summary cell, only one cell (va1 + 2) = 9
        self.set_border(ws_summary,cell_range,wraptext=False,font_weight=True,font_size=14)

        #   set style left side of Total sum
        cell_range = [(len1 + 1), col_ind+1, (len1 + 1) + var2, col_ind+1]  # Cell range of IOS name to make font weight bold
        self.set_border(ws_summary, cell_range, wraptext=False, font_size=14, font_weight=True)

        # set Column width
        sd1 = [9]*(col_ind-1)+[28, 32, 32, 9, 9, 9, 9, 9, 9, 9]
        for sd in range(1,ws_summary.max_column + 1):
            ws_summary.column_dimensions[f"{get_column_letter(sd)}"].width = sd1[sd - 1]

        # set height of the headers
        for itm in length: ws_summary.row_dimensions[itm].height = float(27)

        #  floating number will show up to two decimal point
        num_f = ws_summary.iter_rows(min_row=len0, min_col=col_ind+2, max_row=ws_summary.max_row, max_col=col_ind+2)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        return None

    def set_border(self, ws, cell_range,bd_sty="thin",font_weight=False,bg_color = "ffffff",wraptext = True,font_size = 11):  # Function to set cell properties
        #bd_sty = "thin"
        bd_color = "000000"
        border = Border(
            left=Side(border_style=bd_sty, color=bd_color),
            right=Side(border_style=bd_sty, color=bd_color),
            top=Side(border_style=bd_sty, color=bd_color),
            bottom=Side(border_style=bd_sty, color=bd_color)
        )  # this indentation helps to comment unnecessary command line.
        font = Font(name="Times New Roman",size=font_size,bold=font_weight,color="000000")
        # set font bold,italic are boolen, font color etc
        alignment = Alignment(horizontal="center",vertical="center",wrapText=wraptext)
        patternfill = PatternFill(start_color=bg_color,end_color=bg_color, fill_type="solid")  # set cell background color
        # print("Cell range from cell property method ",cell_range)
        rows = ws.iter_rows(min_row=cell_range[0],min_col=cell_range[1],max_row=cell_range[2],max_col=cell_range[3])  # iter_rows function make a tuple of tuples of cell objects.
        for row in rows:
            for cell in row:
                cell.border = border        # set Border properties
                cell.font = font            # set font styles
                cell.alignment = alignment  # set alignment
                cell.fill = patternfill     # set background color.
        return ws

    def make_IDD_report(self):

        fd1 = open(self.filename, "r+");
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_idd = Workbook();
        sh_list = wb_idd.sheetnames;
        sh_list[0] = "idd raw";
        sh_list = wb_idd.sheetnames;
        ws_idd = wb_idd[sh_list[0]];
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A] * 6 + [B] * 2 + [C]*3 + [D, H] + [O] * 3 + [R]*7  # to be delete column list
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):  # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:  # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            row_list.insert(0, " ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_connect_number = row_list.index("Connect Number")
                index_attempt_number = row_list.index("Attempt Number")
                index_Answer_Number = row_list.index("Answer Number")
                index_Answer_Time = row_list.index("Answer Time")
                row_list.append("ASR")
                row_list.append("ACD")
                row_list.append("CCR")
            else:
                max_ro = itm;
                row_list.append(
                    f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(
                    f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60")
                row_list.append(
                    f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")

            ws_idd.append(row_list);
        fd1.close();
        min_row = ws_idd.min_row
        min_col = ws_idd.min_column
        max_row = ws_idd.max_row
        max_col = ws_idd.max_column
        # col_letter = get_column_letter(max_col)

        ''''
        Setting style to the border and fonts, 
        '''
        cell_range = [min_row, min_col + 1, max_row, max_col]  # set style all sheet.
        self.set_border(ws_idd, cell_range)

        top_cell_range = [1, 2, 1, ws_idd.max_column]  # set style top row/ Header row
        self.set_border(ws_idd, top_cell_range, bd_sty="medium", font_weight=True,
                        bg_color=header_color)  # Border width = medium valid !!

        top_cell_range = [2, ws_idd.max_column - 2, ws_idd.max_row, ws_idd.max_column]  # set style ASR,ACD,CCR
        self.set_border(ws_idd, top_cell_range, bd_sty="thin", font_weight=True, bg_color=side_color)

        rd = ws_idd.row_dimensions[1]  # get dimension for row 3
        rd.height = 48  # value in points, there is no "auto"
        # sd1 = [16,14,9,12,15,15,12,10,10,11,11,10M,12,11,11,13Q,11,11,11]
        sd1 = [4, 16, 14, 13, 15, 15, 12, 10, 11, 11, 12, 12, 13, 12, 12, 13, 13, 13, 12, 9, 9, 9,9]
        for sd in range(1, ws_idd.max_column + 1):
            cd = ws_idd.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        ws_idd.freeze_panes = "A2"  # make freeze before B2
        # ws_idd.insert_cols(1)  # this will insert correctly but little problem with my fix excel formulas

        #  floating number will show up to two decimal point
        num_f = ws_idd.iter_rows(min_row=2, min_col=ws_idd.max_column - 2, max_row=ws_idd.max_row,
                                 max_col=ws_idd.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        # save file with name
        x = time.localtime(os.path.getctime(self.filename))

        date_file = ws_idd.cell(2, 3).value.split("-")
        month_file = month_name[int(date_file[1])]

        if x[3] > 12:        # for morning report it is false.
            time_upto = datetime.strptime(f"{x[3]}", "%H").strftime("%I %p")
            name = f"IDD Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"IDD Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''

        wb_idd.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        xl = Dispatch("Excel.Application")
        wb2 = xl.Workbooks.Open(Filename=path2)
        xl.Visible = False  # speed up process also
        wb2.Save()  # Save and over lap the original file
        wb2.Close(True)
        xl.Quit()

        # Print completion message
        kip_conform.set(f"IDD complete morning / Evening")
        # except Exception: print("File not found \n Or Wrong file selected \nOut file has opened")

        # for Mailing purpose
        upload.nam_temp_idd = name
        upload.nam_temp_idd_t = [date_file[0],month_file,date_file[2],time_upto]
        if auto_mail_enable.get() == 1:
            mail_instance.idd()

        return None


class Open_gui:

    @classmethod
    def menu_bars(cls):
        # Menu make
        menubar = Menu(top)
        file = Menu(menubar,tearoff=False)
        file.add_command(label="Set Header Color", command=Options.headerColor)
        file.add_command(label="Set Side color", command=Options.sideColor)
        file.add_separator()
        file.add_command(label="Exit", command=top.quit)
        menubar.add_cascade(label="File", menu=file)

        tools = Menu(menubar,tearoff=False)
        tools.add_command(label="Digital Clock", command=Tools.d_clock)
        tools.add_command(label="KPI, IOS_ISD, IDD Report", command=Open_gui.kpi_gui)
        global auto_mail_enable
        auto_mail_enable = IntVar()
        tools.add_checkbutton(label="Auto Mail Enable",variable = auto_mail_enable, offvalue=0, onvalue=1)
        menubar.add_cascade(menu=tools, label="tools")

        global mail_instance
        mail_instance = Mail("Md. Habibur Rahman")
        mail_menu = Menu(menubar,tearoff=False)
        mail_menu.add_command(label="create new Mail", command=mail_instance.creat_mail)
        mail_menu.add_command(label="IDD Every Two Hours", command=mail_instance.idd_every_two_hours)
        mail_menu.add_command(label="KPI report", command=mail_instance.kpi)
        mail_menu.add_command(label="IOS ISD report", command=mail_instance.ios_idd)
        mail_menu.add_command(label="IDD report", command=mail_instance.idd)
        menubar.add_cascade(menu=mail_menu, label="Mail")

        noc_name = Menu(menubar,tearoff=False)
        noc_name.add_radiobutton(label="Habib", command=partial(mail_instance.__init__, "Md. Habibur Rahman"))
        noc_name.add_radiobutton(label="Turzo", command=partial(mail_instance.__init__, "Akil Monsur"))
        noc_name.add_radiobutton(label="Turaz", command=partial(mail_instance.__init__, "Tanzil Monsur"))
        noc_name.add_radiobutton(label="Amit", command=partial(mail_instance.__init__, "Amit Roy"))
        noc_name.add_radiobutton(label="Tusher", command=partial(mail_instance.__init__, "Md. Tasnim Rahman Tusher"))
        noc_name.add_radiobutton(label="Obeyddullah", command=partial(mail_instance.__init__, "A.E.M Obeyddullah Siddique"))
        menubar.add_cascade(menu=noc_name, label="Employee")
        top.config(menu=menubar)

        # Excel Header color set
        exists = os.path.isfile(r'data/color/header_color.txt')
        global header_color,side_color
        if exists:
            header_color = open(r'data/color/header_color.txt', "r+").read(6)
        else:
            header_color = "ff0000"

            # Excel side  color set
        exists = os.path.isfile(r'data/color/side_color.txt')
        if exists:
            side_color = open(r'data/color/side_color.txt', "r+").read(6)
        else:
            side_color = "00ff00"
        return None

    @classmethod
    def kpi_gui(cls):
        '''
        Starting point of periodic monitoring
        '''
        top.wm_state('iconic')
        kpi_top = Toplevel()
        kpi_top.resizable(width=False,height=False)

        menubar = Menu(top)
        file = Menu(menubar)
        file.add_command(label="Set Header Color", command=Options.headerColor)
        file.add_command(label="Set Side color", command=Options.sideColor)
        file.add_separator()
        def des():
            kpi_top.destroy()
            top.wm_state('normal')

        file.add_command(label="Exit", command=des)
        menubar.add_cascade(label="File", menu=file)
        kpi_top.config(menu=menubar)

        bg_kpi1 = "#00ff00"
        bg_kpi2 = "#99aa00"
        Label(kpi_top).grid(row=1, column=0)  # gap between two part

        global raw_kpi_in_instance,raw_kpi_out_instance,raw_idd
        raw_kpi_in_instance = upload();
        raw_kpi_out_instance = upload();
        raw_idd = upload();

        frm_kpi = Frame(kpi_top, bd=5, relief="solid", pady=5, padx=5, bg="#ffffff")
        raw_frame = Frame(frm_kpi, bd=5, relief="solid", width=500, bg="#ff009f");
        make_frame = Frame(frm_kpi, bd=5, relief="solid", width=500, bg="#ff009f");

        global raw_in_kpi_file_path,raw_out_kpi_file_path,raw_idd_file_path
        raw_in_kpi_file_path = StringVar();
        raw_in_kpi_file_path.set("Open incoming KPI .csv")
        raw_out_kpi_file_path = StringVar();
        raw_out_kpi_file_path.set("Open outgoing KPI .csv")
        raw_idd_file_path = StringVar();
        raw_idd_file_path.set("Open IDD .csv")

        btn_1 = Button(raw_frame, command=raw_kpi_in_instance.upload_kpi, textvariable=raw_in_kpi_file_path,
                       width=width_1, font="Times 15", bg=bg_kpi1, anchor="w");
        btn_2 = Button(raw_frame, command=raw_kpi_out_instance.upload_kpi_out, textvariable=raw_out_kpi_file_path,
                       width=width_1, font="Times 15", bg=bg_kpi1, anchor="w");
        btn_3 = Button(raw_frame, command=raw_idd.upload_idd, textvariable=raw_idd_file_path, width=width_1,
                       font="Times 15", bg=bg_kpi1, anchor="w");

        Label(frm_kpi, width=width_1, font="Times 6").grid(row=3, column=0)  # gep between row file and make report

        width_2 = round(width_1 / 3) - 1
        btn_kpi = Button(make_frame, text="Make KPI Report", command=raw_kpi_in_instance.make_kpi_report, width=width_2,
                         font=font, bg=bg_kpi2)
        btn_kpi.grid(row=0, column=0);

        btn_ios = Button(make_frame, text="Make IOS IDD Report", command=raw_kpi_out_instance.make_IOS_ISD_report,
                         width=width_2, font=font, bg=bg_kpi2)
        btn_ios.grid(row=0, column=1);

        btn_idd = Button(make_frame, text="Make IDD Report", command=raw_idd.make_IDD_report, width=width_2, font=font,
                         bg=bg_kpi2)
        btn_idd.grid(row=0, column=2);

        btn_1.grid(row=1, column=1);
        btn_2.grid(row=2, column=1);
        btn_3.grid(row=3, column=1);

        raw_frame.grid(row=2, column=0);  # row=2 in main frame "frm_kpi"
        make_frame.grid(row=4, column=0);  # row=4 in main frame "frm_kpi"
        frm_kpi.grid(row=3, column=0)  # this is the main frame for KPI report, inside top

        # show conformation messege
        global kip_conform
        kip_conform = StringVar()
        kip_conform.set(" ")
        Label(frm_kpi, textvariable=kip_conform, font=font, width=width_1,bg="white").grid(row=5, column=0)

        return None

    @classmethod
    def periodic_gui(cls):
        pass


class Tools:

    def __init__(self):
        pass

    @classmethod
    def d_clock(self):
        root = Toplevel()
        clock = Label(root, font="Times 100",bg="white")
        root.resizable(width=False,height=False)
        clock.pack()
        def clock_refresher():
            time_string = time.strftime("%H:%M:%S")
            clock.config(text=time_string)
            clock.after(50, clock_refresher)
        Thread(target=clock_refresher).start()
        return None


class Options:   # Menu make
    def __init__(self):
        pass

    @classmethod
    def headerColor(cls):
        # make required folders
        cwd = os.getcwd()  # reserve current directory
        if not os.path.exists('data'):
            os.mkdir('data')
        os.chdir("data")
        if not os.path.exists('color'):
            os.mkdir('color')
        os.chdir(cwd)  # make current directory as previous

        color_temp = askcolor(color="red", title="select Header color")
        global header_color
        if color_temp[0] != None:
            file_1 = open(r'data/color/header_color.txt', "w+")
            file_1.write(color_temp[1][1:])
            header_color = color_temp[1][1:]
        return None

    @classmethod
    def sideColor(cls):
        # make required folders
        cwd = os.getcwd()  # reserve current directory
        if not os.path.exists('data'):
            os.mkdir('data')
        os.chdir("data")
        if not os.path.exists('color'):
            os.mkdir('color')
        os.chdir(cwd)  # make current directory as previous

        color_temp = askcolor(color="red", title="select side color")
        global side_color
        if color_temp[0] != None:
            file_1 = open(r'data/color/side_color.txt', "w+")
            file_1.write(color_temp[1][1:])
            side_color = color_temp[1][1:]
        return None


class Mail:
    font_size = 14.5

    def __init__(self,name="Md. Habibur Rahman"):

        print(auto_mail_enable.get()) # test "auto_mail_enable" value

        self.header_sty = '<!DOCTYPE html>' \
                '<html lang="en">'  \
                '<head>'    \
                '<meta charset="UTF-8">'    \
                '<title>ICX KPI Report</title>' \
                r'<style> body { font-family: Times New Roman;font-size: ' + f'{Mail.font_size}' + r'px;} </style>' \
                '</head>'   \
                "<body>"

        self.sign = '<br><br><br>' \
                '<b>Best Regards,</b>  <br>' \
                f'{name}<br>' \
                r'Jr.Engineer <br>' \
                r'Network Operation Center(NOC) <br>' \
                f'<img src="{os.getcwd()}\data\image\logo.jpg">' \
                r'<br>E-mail: noc@mmclbd.com<br>' \
                r'Call us: +8801777189722<br>' \
                '</body>' \
                r'</html>'

    def creat_mail(self, to=None, subject=None, cc=None, body=None, bcc=None, attach=None):
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)

        if to != None:
            mail.To = to
        if cc != None:
            mail.cc = cc
        if bcc!= None:
            mail.bcc=bcc
        if subject != None:
            mail.Subject = subject
        if attach != None:
            mail.Attachments.Add(attach)
        if body != None:
            mail.HtmlBody = body
        print("this is create")
        mail.display(False)   # putting True creates mistake
        # mail.send
        return None

    def kpi(self):
        to="raj@mmclbd.com"
        cc="sharma.chandan@mmclbd.com;anirban@mmclbd.com;arif@mmclbd.com;noc@mmclbd.com"

        try:
            attach = os.getcwd() + r'/' + upload.nam_temp_kpi
        except Exception as e:
            print("Error type in KPI : ",e)
            attach = None

        c_time = time.localtime()
        try:
            time_t = upload.nam_temp_kpi_t  # [year, month, day, upto time]
        except Exception: time_t = [c_time[0],month_name[c_time[1]],c_time[2],'']

        subject = f"ICX KPI report on {time_t[2]} {time_t[1]} {time_t[0]}"
        body = self.header_sty + \
            'Dear Vaiya, <br><br>'  \
            f'Please check the attached KPI report of ICX domestic call on {time_t[2]} {time_t[1]} {time_t[0]}{time_t[3]}.' \
            + self.sign

        print(body)
        # obj_body = open("mail_kpi_m.html","r")
        # body = obj_body.read()
        self.creat_mail(to=to,cc=cc,subject=subject,body=body,attach=attach)
        return None

    def ios_idd(self):
        to = "raj@mmclbd.com"
        cc = "sharma.chandan@mmclbd.com;anirban@mmclbd.com;arif@mmclbd.com;noc@mmclbd.com"

        # set attached file name.
        try:
            attach = os.getcwd() + r'/' + upload.nam_temp_ios
        except Exception as e:
            print("Error type in KPI : ", e)
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = upload.nam_temp_ios_t  # [year, month, day, upto time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], '']

        subject = f"IOS & ISD Report on {time_t[2]} {time_t[1]} {time_t[0]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD incoming ( IOS & ISD) report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}.' \
               + self.sign

        print(body)
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def idd(self):
        to = "raj@mmclbd.com"
        cc = "sharma.chandan@mmclbd.com;anirban@mmclbd.com;arif@mmclbd.com;noc@mmclbd.com"

        # set attached file name.
        try:
            attach = os.getcwd() + r'/' + upload.nam_temp_idd
        except Exception as e:
            print("Error type in KPI : ", e)
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = upload.nam_temp_idd_t  # [year, month, day, upto time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], '']

        subject = f"IDD Report on {time_t[2]} {time_t[1]} {time_t[0]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD incoming ( IOS & ISD) report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}.' \
               + self.sign
        print(body)
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def idd_every_two_hours(self):
        to = "raj@mmclbd.com"
        cc = "anirban@mmclbd.com;arif@mmclbd.com;noc@mmclbd.com"

        # set attached file name.
        try:
            attach = os.getcwd() + r'/' + idd_class.filename_mail_2h
        except Exception as e:
            print("Error type in KPI : ", e)
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = idd_class.filename_mail_2h_t  # [year, month, day, upto_time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], f'{time.strftime("%I",c_time)} {time.strftime("%p",c_time)}']
            # Last element of this list will print like "10 AM"

        subject = f"IDD Report on {time_t[2]} {time_t[1]} {time_t[0]} up to {time_t[3]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD Report on  {time_t[2]} {time_t[1]} {time_t[0]} up to {time_t[3]}.' \
               + self.sign
        print(body)
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def all_ios(self):
        to = "raj@mmclbd"
        cc = "anirban@mmclbd;arif@mmclbd;noc@mmclbd"
        bcc = ""
        subject = f"IDD Report on "
        body = self.header_sty + \
                self.sign
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, bcc=bcc)
        return None

if __name__ == "__main__":

    top = Tk();

    top.title("Periodic Monitoring");
    try: top.iconbitmap(default = r'C:\mmcl.ico');
    except Exception:pass;

    top.resizable(width=False,height=False);
    width_of_window = 832;
    height_of_window = 348;
    screen_width = top.winfo_screenwidth();
    screen_height = top.winfo_screenheight();
    x_coordinate = screen_width/2 - width_of_window/2;
    y_coordinate = screen_height/2 - height_of_window/2;
    top.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate));
    top.configure(bg="#ffffff")  #set app background
    bg_1 = "#ffff9f"
    bg_11 = "#ffff4f"
    bg_2 = "#ffa0ff"
    bg_21 = "#ffa08f"
    fg_1 = "#000000"
    width_1 = 70
    font = "Times 15"


    # Menu make
    Open_gui.menu_bars()


    frm = Frame(top,bd=5,relief="solid",pady=5,padx=5,bg="#ffffff")
    frm_1 = Frame(frm,bd=5,relief="solid",width=500,bg="#ff009f")
    frm_2 = Frame(frm,bd=5,relief="solid",width=500,bg="#ff009f")

    ins_text_1 = idd_class();

    var_1 = StringVar(); var_1.set(ins_text_1.filename);
    var_3 = StringVar(); var_3.set("Make IDD Report");

    btn_1 = Button(frm_1, textvariable=var_1, command=ins_text_1.upload, width=width_1, font=font,bg=bg_1,fg=fg_1);
    btn_3 = Button(frm_1, textvariable=var_3, command=ins_text_1.csv_file, width=width_1, font=font,bg=bg_1,fg=fg_1);
    Label(frm_1,text="IDD Every Two hours",bg=bg_11,width=width_1,font=font).grid(row=0,column=0)
    btn_1.grid(row=1,column=0)
    btn_3.grid(row=3,column=0)
    frm_1.grid(row=1,column=0)       # row=1 in main frame "frm"

    Label(frm,width=115).grid(row=0,column=0) # up side blank space      # row=0 in main frame "frm"
    Label(frm,width=115).grid(row=2,column=0) # buttom side blank space  # row=2 in main frame "frm"

    ins_text = ccr_class();
    ins_text2 = ccr_class();
    var1 = StringVar(); var1.set(ins_text.filename);
    var2 = StringVar(); var2.set("CCR Check");
    btn1 = Button(frm_2,textvariable=var1,command=ins_text.upload,width=width_1,font=font,bg=bg_2,fg=fg_1);
    btn2 = Button(frm_2,textvariable=var2,command=ins_text.csv_file,width=width_1,font=font,bg=bg_2, fg=fg_1);
    Label(frm_2,text=" CCR Check ",bg=bg_21,width=width_1,font=font, fg=fg_1).grid(row=0,column=0)
    btn1.grid(row=1,column=0)
    btn2.grid(row=2,column=0)
    frm_2.grid(row=3,column=0)       # row=3 in main frame "frm"

    lab_1 = Label(frm, font="Times 30",bg="white")
    def clock_refresher():
        time_string = time.strftime("%r")
        lab_1.config(text=time_string)
        lab_1.after(100, clock_refresher)
    # clock_refresher()
    t1 = Thread(target=clock_refresher,args=())
    t1.start();

    lab_1.grid(row=4,column=0)
    frm.grid(row=2,column = 0)      # this is the main frame for ccr idd

    top.mainloop();

